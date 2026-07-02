#!/usr/bin/env python3
"""导出 fund_raw_sample 全量数据为 Excel"""
import os, sys, json, time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SUPABASE_URL = "https://tqhtegazxykkqfcpejky.supabase.co"
ANON_KEY = "sb_publishable_iFtMcvav774gqF28gGYQVw_QMmuS-z3"
TABLE = "fund_raw_sample"

# 需要导出的列（按重要性排序）
COLS = [
    "c", "name",
    "t0", "t1",
    "company", "fund_scale", "manage_fee", "custody_fee",
    "total_manage_scale", "holders_count", "risk_level",
    "estab_date", "hold_days", "fund_leverage", "port_duration",
    "nav", "nav_date",
    "r1y", "r2y", "r3y", "r5y",
    "dd1y", "dd2y", "dd3y", "dd5y",
    "sr1y", "sr2y", "sr3y", "sr5y",
    "score_3m", "score_6m", "score_1y", "score_2y", "score_3y", "score_5y", "score_7y", "score_10y",
    "r0w", "r1m", "r3m", "r6m", "ytd",
]

# 列中文名
COL_CN = {
    "c": "基金代码", "name": "基金名称",
    "t0": "大类", "t1": "二级分类",
    "company": "基金公司", "fund_scale": "基金规模(亿)",
    "manage_fee": "管理费/y", "custody_fee": "托管费/y",
    "total_manage_scale": "任职总规模(亿)", "holders_count": "持有人数量",
    "risk_level": "风险等级", "estab_date": "成立日期",
    "hold_days": "持有天数", "fund_leverage": "基金杠杆率", "port_duration": "组合久期",
    "nav": "最新净值", "nav_date": "净值日期",
    "r1y": "近1年收益%", "r2y": "近2年收益%", "r3y": "近3年收益%", "r5y": "近5年收益%",
    "dd1y": "近1年回撤%", "dd2y": "近2年回撤%", "dd3y": "近3年回撤%", "dd5y": "近5年回撤%",
    "sr1y": "近1年夏普", "sr2y": "近2年夏普", "sr3y": "近3年夏普", "sr5y": "近5年夏普",
    "score_3m": "3月评分", "score_6m": "6月评分", "score_1y": "1年评分",
    "score_2y": "2年评分", "score_3y": "3年评分", "score_5y": "5年评分",
    "score_7y": "7年评分", "score_10y": "10年评分",
    "r0w": "本周收益%", "r1m": "近1月收益%", "r3m": "近3月收益%", "r6m": "近6月收益%", "ytd": "年初至今%",
}

def fetch_all():
    """分页拉取全量数据"""
    all_rows = []
    offset = 0
    limit = 1000
    total = None

    print(f"开始拉取 {TABLE}...")

    headers = {
        "apikey": ANON_KEY,
        "Authorization": f"Bearer {ANON_KEY}",
        "Accept": "application/json",
    }

    while True:
        url = f"{SUPABASE_URL}/rest/v1/{TABLE}"
        params = {
            "select": ",".join(COLS),
            "offset": str(offset),
            "limit": str(limit),
            "order": "c.asc",
        }

        resp = requests.get(url, headers=headers, params=params, timeout=60)
        if resp.status_code != 200:
            print(f"  错误: HTTP {resp.status_code} at offset {offset}: {resp.text[:200]}")
            break

        rows = resp.json()
        if not rows:
            break

        all_rows.extend(rows)
        offset += len(rows)

        if total is None and offset >= 1000:
            # 用 count 头获取总数
            try:
                cr = requests.get(f"{SUPABASE_URL}/rest/v1/{TABLE}?select=c&limit=0",
                                  headers={**headers, "Prefer": "count=exact"}, timeout=30)
                range_str = cr.headers.get("content-range", "")
                if "/" in range_str:
                    total = int(range_str.split("/")[-1])
                    print(f"  总行数: {total}")
            except:
                pass

        if len(rows) < limit:
            break

        if offset % 5000 == 0:
            print(f"  已拉取: {offset}/{total or '?'}")
        time.sleep(0.05)

    print(f"  拉取完成: {len(all_rows)} 行")
    return all_rows

def export_excel(rows, filepath):
    """导出为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "fund_raw_sample"

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1d70b8", end_color="1d70b8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 表头
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=COL_CN.get(col, col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(COLS, 1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

        if ri % 5000 == 0:
            print(f"  写入: {ri-1}/{len(rows)}")

    # 列宽
    col_widths = {
        "c": 10, "name": 24, "t0": 10, "t1": 16,
        "company": 16, "fund_scale": 14, "manage_fee": 12, "custody_fee": 12,
        "total_manage_scale": 16, "holders_count": 14,
        "risk_level": 10, "estab_date": 12,
        "hold_days": 10, "fund_leverage": 12, "port_duration": 12,
        "nav": 10, "nav_date": 12,
    }
    for ci, col in enumerate(COLS, 1):
        width = col_widths.get(col, 12)
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width

    # 冻结首行
    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"  保存: {filepath}")

def main():
    rows = fetch_all()
    if not rows:
        print("没有数据，退出")
        return

    outdir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
    os.makedirs(outdir, exist_ok=True)

    filepath = os.path.join(outdir, "fund_raw_sample_full.xlsx")
    export_excel(rows, filepath)

    # 统计本次新增填充的列
    print(f"\n=== 数据填充统计 ({len(rows)} 行) ===")
    newly_filled = [
        "company", "fund_scale", "manage_fee", "custody_fee",
        "total_manage_scale", "holders_count", "risk_level", "estab_date",
        "hold_days", "fund_leverage", "port_duration",
    ]
    for col in newly_filled:
        filled = sum(1 for r in rows if r.get(col) is not None and r.get(col) != "")
        pct = filled / len(rows) * 100 if rows else 0
        print(f"  {COL_CN.get(col, col):14s}: {filled:>6}/{len(rows)} ({pct:.1f}%)")

    print(f"\n文件: {filepath}")
    print("完成!")

if __name__ == "__main__":
    main()
