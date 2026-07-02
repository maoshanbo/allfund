#!/usr/bin/env python3
"""导出完整 fund_combined 表为 Excel（支持本地和 CI 环境）"""

import os, sys, argparse, requests, csv, time

# 兼容两种环境变量命名（本地: VITE_SUPABASE_*, CI: SUPABASE_*）
SUPABASE_URL = os.environ.get('VITE_SUPABASE_URL') or os.environ.get('SUPABASE_URL') or 'https://tqhtegazxykkqfcpejky.supabase.co'
ANON_KEY = os.environ.get('VITE_SUPABASE_ANON_KEY') or os.environ.get('SUPABASE_ANON_KEY') or 'sb_publishable_iFtMcvav774gqF28gGYQVw_QMmuS-z3'
REST = f'{SUPABASE_URL}/rest/v1/fund_combined'

# 命令行参数
parser = argparse.ArgumentParser(description='导出 fund_combined 为 Excel')
parser.add_argument('--output', default=None, help='Excel 输出路径（默认: exports/fund_combined.xlsx）')
parser.add_argument('--csv', default=None, help='CSV 输出路径（默认: exports/fund_combined_full.csv）')
args = parser.parse_args()

# 输出路径：脚本所在目录的相对路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
DEFAULT_XLSX = os.path.join(PROJECT_DIR, 'exports', 'fund_combined.xlsx')
DEFAULT_CSV = os.path.join(PROJECT_DIR, 'exports', 'fund_combined_full.csv')

xlsx_output = args.output or DEFAULT_XLSX
csv_output = args.csv or DEFAULT_CSV

# 确保输出目录存在
os.makedirs(os.path.dirname(xlsx_output), exist_ok=True)

HEADERS = {
    'apikey': ANON_KEY,
    'Authorization': f'Bearer {ANON_KEY}',
}

# 实际存在于 fund_combined 的列
COLS = [
    'c', 'name', 't0', 't1', 'company', 'fund_scale', 'risk_level', 'manage_fee',
    'ytd', 'r1y', 'r3y', 'r5y',
    'dd1y', 'sr1y',
    'holders_count', 'total_manage_scale',
    'score_grade',
    'k_all', 'k0w', 'k1m', 'k3m', 'k6m', 'k1', 'k2', 'k3', 'k5', 'k7', 'k10',
]
COL_CN = {
    'c': '基金代码', 'name': '基金名称', 't0': '一级分类', 't1': '二级分类',
    'company': '基金公司', 'fund_scale': '基金规模(亿)', 'risk_level': '风险等级', 'manage_fee': '管理费/y',
    'ytd': 'YTD(%)', 'r1y': '近1年(%)', 'r3y': '近3年(%)', 'r5y': '近5年(%)',
    'dd1y': '1年最大回撤(%)', 'sr1y': '1年夏普比率',
    'holders_count': '持有人数', 'total_manage_scale': '总管理规模(亿)',
    'score_grade': '评级',
    'k_all': '综合评分', 'k0w': '本周评分', 'k1m': '1月评分',
    'k3m': '3月评分', 'k6m': '6月评分', 'k1': '1年评分', 'k2': '2年评分',
    'k3': '3年评分', 'k5': '5年评分', 'k7': '7年评分', 'k10': '10年评分',
}

print('开始导出 fund_combined 全量数据...')

# 分页拉取
all_rows = []
offset = 0
page_size = 1000

while True:
    cols_str = ','.join(COLS)
    url = f'{REST}?select={cols_str}&limit={page_size}&offset={offset}'
    r = requests.get(url, headers=HEADERS, timeout=60)
    if r.status_code != 200:
        print(f'  HTTP {r.status_code} at offset {offset}: {r.text[:200]}')
        break
    batch = r.json()
    if not batch:
        break
    all_rows.extend(batch)
    print(f'  已拉取 {len(all_rows)} 行...')
    offset += page_size
    time.sleep(0.1)

print(f'共拉取 {len(all_rows)} 行')

# 写入 CSV
with open(csv_output, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow([COL_CN.get(c, c) for c in COLS])
    for row in all_rows:
        vals = []
        for c in COLS:
            v = row.get(c, '')
            if v is None:
                v = ''
            elif isinstance(v, float):
                v = round(v, 2)
            vals.append(v)
        w.writerow(vals)

print(f'✓ CSV: {csv_output}')

# 转 Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    print('生成 Excel...')
    wb = Workbook()
    ws = wb.active
    ws.title = 'fund_combined'

    # 表头
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1D70B8', end_color='1D70B8', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    for ci, c in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=COL_CN.get(c, c))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    blue_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    for ri, row in enumerate(all_rows, 2):
        for ci, c in enumerate(COLS, 1):
            v = row.get(c, '')
            if v is None:
                v = ''
            elif isinstance(v, float):
                v = round(v, 2)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # 按评级着色
        grade = row.get('score_grade', '')
        if grade == 'green':
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = green_fill
        elif grade == 'blue':
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = blue_fill
        elif grade == 'orange':
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = orange_fill

    # 列宽
    col_widths = {
        'A': 10, 'B': 28, 'C': 10, 'D': 18, 'E': 16, 'F': 14, 'G': 8, 'H': 8,
    }
    for ci in range(9, len(COLS) + 1):
        col_widths[get_column_letter(ci)] = 12

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    xlsx_path = xlsx_output
    wb.save(xlsx_path)
    print(f'✓ Excel: {xlsx_path} ({len(all_rows)} 行)')

except ImportError:
    xlsx_path = csv_path
    print('⚠ openpyxl 未安装，仅生成 CSV')

print('完成!')
