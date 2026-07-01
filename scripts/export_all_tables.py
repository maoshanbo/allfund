#!/usr/bin/env python3
"""
导出 Supabase 所有 public 表为 Excel 文件到 public/downloads/ 目录
用法: python scripts/export_all_tables.py [--output-dir public/downloads]
"""
import os, sys, json, requests
from datetime import datetime

# ===== 配置 =====
SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('VITE_SUPABASE_URL') or 'https://tqhtegazxykkqfcpejky.supabase.co'
SUPABASE_KEY = os.environ.get('SUPABASE_ANON_KEY') or os.environ.get('VITE_SUPABASE_ANON_KEY') or 'sb_publishable_iFtMcvav774gqF28gGYQVw_QMmuS-z3'

# 判断是否 CI 环境
IS_CI = os.environ.get('CI') == 'true' or os.environ.get('GITHUB_ACTIONS') == 'true'
if IS_CI:
    import subprocess
    result = subprocess.run(['pip', 'install', 'openpyxl'], capture_output=True, text=True)
    print(f"[CI] pip install openpyxl: {result.returncode}")
else:
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl requests")
        sys.exit(1)

import openpyxl
from openpyxl.utils import get_column_letter

# 输出目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'public', 'downloads')
if '--output-dir' in sys.argv:
    idx = sys.argv.index('--output-dir')
    OUTPUT_DIR = sys.argv[idx + 1]
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
}

# ===== 表定义 =====
TABLES = {
    'fund_combined': {
        'name': '基金综合数据表',
        'desc': '基金分类(t0/t1)、公司/规模/费率、收益(ytd/r1y/r3y/r5y)、风险指标(dd1y/sr1y)、持有人数、评分(k_all/score_grade/k0w~k10) — 所有数据核心合并表，19+ 周期评分全覆盖',
        'source': '天天基金 FundGuideapi（收益率/分类）+ pingzhongdata（回撤/夏普/风险评级）+ rankhandler（货币基金收益）+ fundf10（公司/规模/费率）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': True,
    },
    'fund_scores': {
        'name': '基金评分表（完整版）',
        'desc': '每日更新的核心数据表：基金代码/名称/基金经理/管理人/分类/规模/费率 → 阶段收益(ytd/r0w~r10y/return_all) → 阶段回撤(dd1y~dd5y) → 阶段夏普(sr1y~sr5y) → 基金评分(k0w~k_all/score_grade)。按以上顺序排列。',
        'source': 'FundGuideapi（收益率/分类）+ pingzhongdata（回撤/夏普/基金经理）+ fund_combined（公司/规模/费率）+ rankhandler（货币基金/成立以来收益）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': True,
    },
    'fund_quarterly_scores': {
        'name': '季度评分表',
        'desc': '基于季报数据的各时间窗口评分（score_3m/6m/1y/2y/3y/5y/7y/10y），含原始 quarterly_data JSON',
        'source': 'pingzhongdata 每日净值 → 季度收益/回撤/夏普计算 → 全市场排名 → 多周期均值评分',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'fund_scores_meta': {
        'name': '评分元数据表',
        'desc': '评分更新时间、基金总数、有评分数、净值日期等元信息',
        'source': 'Supabase 内部自动记录',
        'update': '每次评分计算完成后自动更新',
        'scoring': False,
    },
    'config': {
        'name': '配置表',
        'desc': '全站配置项（键值对，含 meta、tsq 时间戳）',
        'source': '手动维护',
        'update': '按需手动更新',
        'scoring': False,
    },
    'macro_history': {
        'name': '宏观历史数据表',
        'desc': '中国10年国债(cn10y)、美国10年国债(us10y)、Shibor、CPI、M2 的历史数据，覆盖 1996-至今',
        'source': 'akshare 开源 Python 库（自动采集公开宏观数据）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'index_pe_history': {
        'name': '指数PE历史表',
        'desc': '沪深300等指数的 PE/PB 历史估值数据',
        'source': '腾讯行情 qt.gtimg.cn + 蛋卷基金 danjuanfunds.com',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'site_stats': {
        'name': '站点统计表',
        'desc': '网站访问量等统计指标',
        'source': 'EdgeOne Pages 边缘函数自动记录',
        'update': '实时更新',
        'scoring': False,
    },
    'tougu_products': {
        'name': '投顾产品表',
        'desc': '天天基金/华宝/盈米/新浪仓石四来源的基金投顾产品，含收益率、最大回撤、标签分类',
        'source': '天天基金投顾页面 + 华宝/盈米/新浪仓石官方数据',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'user_portfolios': {
        'name': '用户组合表',
        'desc': '用户自建基金组合数据（portfolio_data JSON），关联用户 ID',
        'source': '用户通过 ALLFUND.CN 网站自行创建',
        'update': '用户操作时实时更新',
        'scoring': False,
    },
    'user_profiles': {
        'name': '用户档案表',
        'desc': '用户注册信息、登录次数、最后登录时间',
        'source': '用户注册时填写',
        'update': '用户操作时实时更新',
        'scoring': False,
    },
}

# 敏感表 / 含用户数据 — 仅在"我的"页面已登录时可见下载
SENSITIVE_TABLES = {'user_portfolios', 'user_profiles'}

# ===== 导出 =====
def get_table_data(table_name):
    """分页获取表数据"""
    all_rows = []
    offset = 0
    limit = 1000  # Supabase REST API default max rows per request
    
    while True:
        resp = requests.get(
            f'{SUPABASE_URL}/rest/v1/{table_name}?select=*&limit={limit}&offset={offset}',
            headers=HEADERS, timeout=60
        )
        if resp.status_code != 200:
            print(f'  ⚠️ {table_name}: HTTP {resp.status_code}')
            break
        rows = resp.json()
        if not rows:
            break
        all_rows.extend(rows)
        offset += len(rows)
        print(f'  📥 {table_name}: {offset} rows...')
        if len(rows) < limit:
            break
    
    return all_rows

# ===== 列顺序定义（按用户要求的展示顺序） =====
FUND_SCORES_COL_ORDER = [
    'c','n','fund_manager','company','t0','t1','fund_scale','manage_fee',
    'ytd','r0w','r1m','r3m','r1y','r3y','r5y','r7y','r10y','return_all',
    'dd1y','dd2y','dd3y','dd5y',
    'sr1y','sr2y','sr3y','sr5y',
    'k0w','k1m','k3m','k6m','k1','k2','k3','k5','k_all','score_grade',
]
COLUMN_ORDER = {
    'fund_scores': FUND_SCORES_COL_ORDER,
}

def export_to_excel(table_name, rows, output_path):
    """导出为 Excel，含数据说明 sheet"""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    
    if not rows:
        ws.append(['(空表)'])
    else:
        # 列名：优先使用定义的顺序，否则用自然顺序
        col_order = COLUMN_ORDER.get(table_name)
        if col_order:
            # 只保留实际存在的列
            columns = [c for c in col_order if c in rows[0]]
            # 追加未在顺序中定义的新列
            extra = [c for c in rows[0].keys() if c not in columns]
            columns.extend(extra)
        else:
            columns = list(rows[0].keys())
        ws.append(columns)
        
        # 加粗表头
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # 数据行
        import json as _json
        def safe_val(v):
            if v is None: return ''
            if isinstance(v, (int, float, str, bool)): return v
            try: return _json.dumps(v, ensure_ascii=False)
            except: return str(v)
        
        for row in rows:
            ws.append([safe_val(row.get(col, '')) for col in columns])
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    # ===== 添加"数据说明" sheet =====
    ws_meta = wb.create_sheet('数据说明')
    meta = TABLES.get(table_name, {})
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=14, color='1d70b8')
    label_font = Font(name='微软雅黑', bold=True, size=11)
    value_font = Font(name='微软雅黑', size=11)
    note_font = Font(name='微软雅黑', size=10, color='666666')
    wrap_align = Alignment(wrap_text=True, vertical='top')
    
    ws_meta.column_dimensions['A'].width = 18
    ws_meta.column_dimensions['B'].width = 80
    
    row_idx = 1
    
    # 标题
    ws_meta.cell(row=row_idx, column=1, value='ALLFUND.CN 数据说明').font = header_font
    ws_meta.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    row_idx += 2
    
    # 基本信息
    info_rows = [
        ('表名', table_name),
        ('中文名称', meta.get('name', '')),
        ('说明', meta.get('desc', '')),
        ('数据来源', meta.get('source', '')),
        ('更新频率', meta.get('update', '')),
        ('导出时间', export_time),
        ('行数', len(rows)),
    ]
    for label, value in info_rows:
        ws_meta.cell(row=row_idx, column=1, value=label).font = label_font
        c = ws_meta.cell(row=row_idx, column=2, value=str(value))
        c.font = value_font
        c.alignment = wrap_align
        row_idx += 1
    
    # 评分表增加评分说明
    if meta.get('scoring'):
        row_idx += 1
        ws_meta.cell(row=row_idx, column=1, value='评分方法 (V7)').font = Font(name='微软雅黑', bold=True, size=13, color='1d70b8')
        ws_meta.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        row_idx += 1
        
        scoring_notes = [
            ('算法版本', 'V7 — 收益 50% + 回撤 25% + 夏普 25%'),
            ('数据来源', 'FundGuideapi（阶段收益率 r0w~r5y）+ pingzhongdata（回撤 dd1y~dd5y、夏普 sr1y~sr5y）+ rankhandler（货币基金）'),
            ('百分位排名', '全市场基金按各指标降序排名，percentile = (1 - rank/(N-1)) × 100，范围 0~100'),
            ('短周期 k0w/k1m/k3m/k6m', '仅用收益率百分位排名：k_short = ret_percentile'),
            ('长周期 k1/k2/k3/k5', '三维度加权：k_long = 50% × ret_percentile + 25% × dd_percentile + 25% × sr_percentile'),
            ('综合评分 k_all', 'k_all = (k0w×5 + k1m×5 + k3m×10 + k6m×15 + k1×20 + k2×20 + k3×15 + k5×10) / total_weight（仅有效周期参与）'),
            ('评级 score_grade', '按 k_all 百分位分级：green(前20%) > blue(20%-50%) > orange(后50%) > gray(无数据)'),
            ('回撤计算', 'dd_max = -max((peak - nav[i]) / peak) × 100，负数百分比（如 -15.23 表示最大回撤 15.23%）'),
            ('夏普计算', 'Sharpe = (E[Rdaily] - Rf) / σdaily × √250，无风险利率 Rf = 2%/年 = 0.02/250 = 0.00008'),
            ('周期权重', 'k0w:5%, k1m:5%, k3m:10%, k6m:15%, k1:20%, k2:20%, k3:15%, k5:10%（总和=100，天然归一化）'),
        ]
        
        for label, value in scoring_notes:
            ws_meta.cell(row=row_idx, column=1, value=label).font = label_font
            c = ws_meta.cell(row=row_idx, column=2, value=str(value))
            c.font = value_font
            c.alignment = wrap_align
            row_idx += 1
    
    # 免责声明
    row_idx += 1
    ws_meta.cell(row=row_idx, column=1, value='免责声明').font = Font(name='微软雅黑', bold=True, size=11, color='999999')
    row_idx += 1
    disclaimer = (
        '本数据由 ALLFUND.CN 通过公开数据接口自动采集和计算，仅供参考，不构成任何投资建议。'
        '数据可能存在延迟或误差，请以天天基金等官方平台实时数据为准。'
        '投资有风险，入市需谨慎。'
    )
    c = ws_meta.cell(row=row_idx, column=1, value=disclaimer)
    c.font = note_font
    c.alignment = Alignment(wrap_text=True)
    ws_meta.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    
    wb.save(output_path)
    return len(rows)

def main():
    print(f'📊 导出 allfund 数据库全部表到 {OUTPUT_DIR}/')
    print(f'⏰ {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
    
    results = {}
    total_size = 0
    
    for table_name in sorted(TABLES.keys()):
        print(f'⬇️ 导出 {table_name} ...')
        rows = get_table_data(table_name)
        
        output_path = os.path.join(OUTPUT_DIR, f'{table_name}.xlsx')
        count = export_to_excel(table_name, rows, output_path)
        
        size_mb = os.path.getsize(output_path) / (1024 * 1024)
        total_size += size_mb
        results[table_name] = {'rows': count, 'size_mb': size_mb, 'path': f'/downloads/{table_name}.xlsx'}
        print(f'  ✅ {table_name}: {count} rows, {size_mb:.2f}MB → {table_name}.xlsx')
    
    # 保存 JSON 索引文件（供前端读取）
    index_path = os.path.join(OUTPUT_DIR, 'index.json')
    with open(index_path, 'w') as f:
        json.dump({
            'updated_at': datetime.now().isoformat(),
            'tables': results,
        }, f, ensure_ascii=False, indent=2)
    
    print(f'\n✅ 全部完成！{len(results)} 张表，总大小 {total_size:.2f}MB')
    print(f'📋 索引文件: {index_path}')

if __name__ == '__main__':
    main()
