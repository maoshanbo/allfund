#!/usr/bin/env python3
"""导出 fund_raw_sample 全量数据为 Excel"""
import os, sys, requests, json, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

SUPABASE_URL = "https://tqhtegazxykkqfcpejky.supabase.co"
SUPABASE_ANON_KEY = "sb_publishable_iFtMcvav774gqF28gGYQVw_QMmuS-z3"
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "exports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def fetch_all():
    headers = {"apikey": SUPABASE_ANON_KEY}
    all_data = []
    offset = 0
    limit = 1000
    total = None
    
    # First get count
    resp = requests.get(f"{SUPABASE_URL}/rest/v1/fund_raw_sample?select=c&limit=1", 
                        headers={**headers, "Prefer": "count=exact"})
    total = 0
    if "content-range" in resp.headers:
        # Format: "0-0/19456" or "*/19456"
        parts = resp.headers["content-range"].split("/")
        if len(parts) == 2 and parts[1] != "*":
            total = int(parts[1])
    print(f"总记录数: {total}")
    
    while len(all_data) < total:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/fund_raw_sample?"
            f"select=c,name,t0,t1,company,fund_scale,risk_level,manage_fee,ytd,r1y,r3y,holders_count,total_manage_scale,nav,nav_date"
            f"&order=c&limit={limit}&offset={offset}",
            headers=headers
        )
        batch = resp.json()
        if not batch:
            break
        all_data.extend(batch)
        offset += limit
        if len(all_data) % 5000 < limit:
            print(f"  已读取: {len(all_data)}/{total}")
    return all_data

def export(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "fund_raw_sample"
    
    # Header styles
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1D70B8", end_color="1D70B8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Column definitions
    columns = [
        ("c", "基金代码", 12),
        ("name", "基金名称", 30),
        ("t0", "一级分类", 10),
        ("t1", "二级分类", 22),
        ("company", "基金公司", 22),
        ("fund_scale", "基金规模(亿)", 14),
        ("risk_level", "风险等级", 10),
        ("manage_fee", "管理费/y", 10),
        ("ytd", "今年以来%", 10),
        ("r1y", "近1年%", 10),
        ("r3y", "近3年%", 10),
        ("holders_count", "持有人数", 14),
        ("total_manage_scale", "任职总规模(亿)", 14),
        ("nav", "最新净值", 10),
        ("nav_date", "净值日期", 12),
    ]
    
    # Write headers
    for col_idx, (_, title, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = width
    
    # Fix column widths
    for col_idx, (_, _, width) in enumerate(columns, 1):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = width
    
    # Write data
    for row_idx, fund in enumerate(data, 2):
        for col_idx, (field, _, _) in enumerate(columns, 1):
            value = fund.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Microsoft YaHei", size=10)
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{chr(64+len(columns))}1"
    
    # Add summary sheet
    ws2 = wb.create_sheet("分类统计")
    from collections import Counter
    t0_dist = Counter(r["t0"] for r in data)
    
    ws2.cell(row=1, column=1, value="一级分类").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="数量").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 12
    
    for i, (t0, count) in enumerate(sorted(t0_dist.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=i, column=1, value=t0).font = Font(name="Microsoft YaHei", size=10)
        ws2.cell(row=i, column=2, value=count).font = Font(name="Microsoft YaHei", size=10)
    
    # Save
    path = os.path.join(OUTPUT_DIR, "fund_raw_sample_classification.xlsx")
    wb.save(path)
    print(f"\n✅ 已导出: {path}")
    print(f"   总行数: {len(data)}")
    return path

def main():
    print("导出 fund_raw_sample ...")
    data = fetch_all()
    path = export(data)

if __name__ == "__main__":
    main()
